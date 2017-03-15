from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkinter import font
from inventory import inventory
from time import sleep
from drivers import drivers
from trucks import trucks
from datetime import date
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font



class AutocompleteEntry(Entry):
    """
    Subclass of tkinter.Entry that features autocompletion.

    To enable autocompletion use set_completion_list(list) to define a list of
    possible strings to hit.
    To cycle through hits use down and up arrow keys.

    Based off of a class of the same name by Mitja Martini, Russell Adams, and
    Dominic Kexel.
    """

    def set_completion_list(self, completion_list):
        self._competion_list = sorted(completion_list, key=str.lower)
        self._hits = []
        self._hit_index = 0
        self.position = 0
        self.bind('<KeyRelease>', self.handle_keyrelease)

    def autocomplete(self, delta=0):
        """
        Autocomplete the Entry, delta may be 0/1/-1 to cycle through possible
        hits.
        """
        # Need to delete selection otherwise we would fix the current position.
        if delta:
            self.delete(self.position, END)
        else:
            self.position = len(self.get())
        _hits = []
        for element in self._completion_list:
            # Match case insensitivity.
            if element.lower().startswith(self.get().lower()):
                _hits.append(element)
        # If we have a new hit list, keep this in mind.
        if _hits != self._hits:
            self._hits_index = 0
            self._hits = _hits
        # Only allow cycling if we are in a known hit list.
        if _hits == self.hits and self._hits:
            self._hit_index = (self._hit_index + delta) % len(self._hits)
        # Perform the autocompletion
        if self._hits:
            self.delete(0, END)
            self.insert(0, self._hits[self._hit_index])
            self.insert_range(self.position, END)

    def handle_keyrelease(self, event):
        """
        Event handler for they keyrelease event on this widget.
        """
        if event.keysym == 'BackSpace':
            self.delete(self.index(INSERT), END)
            self.position = self.index(END)
        if event.keysym == 'Left':
            if self.position < self.index(END):
                self.delete(self.position, END)
            else:
                self.position = self.position - 1  # Delete one character.
                self.delete(self.position, END)
        if event.keysym == 'Right':
            self.position == self.index(END)  # Go to end (no selection)
        if event.keysym == 'Down':
            self.autocomplete(1)  # Cycle to next hit.
        if event.keysym == 'Up':
            self.autocomplete(-1)  # Cycle to previous hit.
        if len(event.keysym) == 1:
            self.autocomplete()


class AutocompleteCombobox(ttk.Combobox):
    def set_completion_list(self, completion_list):
        """
        Use our completion list as our drop down selection menu, arrows move
        through menu.
        """
        self._completion_list = sorted(completion_list, key=str.lower)
        self._hits = []
        self._hit_index = 0
        self.position = 0
        self.bind('<KeyRelease>', self.handle_keyrelease)
        self['values'] = self._completion_list  # Setup our popup menu.

    def autocomplete(self, delta=0):
        """
        Autocomplete the Combobox, delta may be 0/1/-1 to cycle through
        possible hits.
        """
        # Need to delete selection otherwise we would fix current position.
        if delta:
            self.delete(self.position, END)
        else:
            self.position = len(self.get())
        _hits = []
        for element in self._completion_list:
            if element.lower().startswith(self.get().lower()):
                _hits.append(element)
        # If we have a new hit list, keep this in mind.
        if _hits != self._hits:
            self._hit_index = 0
            self._hits = _hits
        # Only allow cycling if we are in a known hit list.
        if _hits == self._hits and self._hits:
            self._hit_index = (self._hit_index + delta) % len(self._hits)
        # Perform the autocompletion
        if self._hits:
            self.delete(0, END)
            self.insert(0, self._hits[self._hit_index])
            self.select_range(self.position, END)
            try:
                desc.set(inventory[item.get()])
            except:
                desc.set('')

    def handle_keyrelease(self, event):
        """
        Event handler for the keyrelease event on this widget.
        """
        if event.keysym == 'BackSpace':
            self.delete(self.index(INSERT), END)
            self.position = self.index(END)
        if event.keysym == 'Left':
            if self.position < self.index(END):  # Delete the selection.
                self.delete(self.position, END)
            else:
                self.position = self.position - 1  # Delete one character.
                self.delete(self.position, END)
        if event.keysym == 'Right' or event.keysym == 'KP_Enter':
            self.position == self.index(END)  # Go to end (no selection)
        if len(event.keysym) == 1:
            self.autocomplete()


def update_item(event):
    desc.set(inventory[item.get()])


def submit(_=''):
    q = quantity.get(1.0, END)[:-1]
    try:
        if int(float(q)) != float(q):
            raise
    except:
        desc.set("Quantity must be an integer value.")
        return 'break'
    else:
        q = int(float(q))
        today = date.today().strftime("%m/%d/%Y")
        if driver.get() == "" or truck.get() == "" or item.get() == "":
            desc.set("Must have driver, truck, and item.")
        else:
            try:
                book = load_workbook('S:\Requests.xlsx')
            except FileNotFoundError:
                book = Workbook()
            sheet = book.active
            heading = ["Date", "Driver", "Truck", "Item", "Quantity"]
            for c in range(1, len(heading)+1):
                _cell = sheet.cell(row=1, column=c, value=heading[c-1])
                _cell.font = Font(bold=True)
            row = [today, driver.get(), truck.get(), item.get(), q]
            active_row = sheet.max_row+1
            for c in range(1, len(row)+1):
                _cell = sheet.cell(row=active_row, column=c, value=row[c-1])
            desc.set("Succesfully submitted %i of %s." % (q, item.get()))
            item.set("")
            quantity.delete(1.0, END)
            book.save('S:\Requests.xlsx')
        return 'break'


inventory_list = list(inventory.keys())
root = Tk()
default_font = font.nametofont("TkDefaultFont")
default_font.configure(size=14)
root.option_add("*Font", default_font)
Grid.rowconfigure(root, 0, weight=0, pad=3)
Grid.rowconfigure(root, 1, weight=0, pad=3)
Grid.rowconfigure(root, 2, weight=0, pad=3)
# Grid.columnconfigure(root, 0, weight=0)
root.geometry("420x350+30+30")
driverlabel = Label(root, text="Driver:")
driverlabel.grid(row=0, column=0, sticky=N)
driver = StringVar()
driverlist = AutocompleteCombobox(root, textvariable=driver, values=drivers)
driverlist.set_completion_list(drivers)
driverlist.grid(row=0, column=1, sticky=N)
trucklabel = Label(root, text="Truck:")
trucklabel.grid(row=1, column=0, sticky=N)
truck = StringVar()
trucklist = AutocompleteCombobox(root, textvariable=truck, values=trucks)
trucklist.set_completion_list(trucks)
trucklist.grid(row=1, column=1, sticky=N)
combolabel = Label(root, text="Item:")
combolabel.grid(row=2, column=0, sticky=N)
item = StringVar()
combo = AutocompleteCombobox(root, textvariable=item, width=20)
desc = StringVar()
combo.bind('<<ComboboxSelected>>', update_item)
combo.set_completion_list(inventory_list)
combo.grid(row=2, column=1, sticky=N)
description = Label(root, textvariable=desc, wraplength=300, width=30)
description.grid(row=3, column=0, columnspan=2, sticky=N)
quantitylabel = Label(root, text="Quantity")
quantitylabel.grid(row=1, column=2, sticky=N)
quantity = Text(root, height=1, width=5)
quantity.grid(row=2, column=2, sticky=W, padx=10)
quantity.bind('<Return>', submit)
submitbutton = Button(root, text="Submit", command=submit)
submitbutton.grid(row=3, column=2, sticky=N)
combo.focus_set()
root.mainloop()
